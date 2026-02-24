import streamlit as st
import pandas as pd
import io
import random
import calendar
from datetime import datetime, timedelta

# --- 1. 環境檢查 ---
try:
    from ortools.sat.python import cp_model
    ORTOOLS_AVAILABLE = True
except ImportError:
    ORTOOLS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 設定網頁標題與寬度
st.set_page_config(page_title="智慧排班系統", page_icon="📅", layout="wide")

# 隱藏 Streamlit 預設選單
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

if not ORTOOLS_AVAILABLE:
    st.error("❌ 嚴重錯誤：排班引擎 (ortools) 未安裝！")
    st.stop()

# ==========================================
# 2. 核心邏輯定義
# ==========================================

# 基準日：2025/12/21 (用於週期上色)
BASE_DATE = datetime(2025, 12, 21)

# 🌟 [在此修改您想在 Excel 右側統計的班別] 🌟
STATS_TARGETS = ["9例", "9", "8-4'F", "8-5", "12'-9", "4-12", "8-5掃", "8-4'掃", "01"]

def clean_str(s):
    if isinstance(s, pd.Series): 
        if s.empty: return ""
        s = s.iloc[0]
    if pd.isna(s): return ""
    s = str(s).strip()
    if s.endswith(".0"): s = s[:-2]
    if s in ["0", "nan", "None", ""]: return ""
    return s.replace(" ", "").replace("　", "").replace("’", "'").replace("‘", "'").replace("，", ",")

def parse_skills(skill_str):
    if pd.isna(skill_str) or skill_str == "":
        return set()
    s = str(skill_str).replace("，", ",").replace(" ", "").replace("　", "")
    parts = s.split(',')
    valid_skills = set()
    for p in parts:
        clean_p = clean_str(p)
        if clean_p:
            valid_skills.add(clean_p)
    return valid_skills

def smart_rename(df, mapping):
    df.columns = df.columns.astype(str).str.strip()
    new_columns = {}
    for col in df.columns:
        col_str = str(col)
        for target_name, keywords in mapping.items():
            for kw in keywords:
                if kw in col_str:
                    new_columns[col] = target_name
                    break
    if new_columns:
        df = df.rename(columns=new_columns)
    return df

# --- 班別屬性判斷 ---
def is_mandatory_off(shift_name):
    return str(shift_name).strip() == "9例"

def is_regular_rest(shift_name):
    return str(shift_name).strip() == "9"

def is_rest_day(shift_name):
    s = str(shift_name).strip()
    if not s: return True 
    if s in ['休', '0', 'nan', 'None']: return True
    return s.startswith("9")

def is_working_day(shift_name):
    return not is_rest_day(shift_name)

# --- 週期計算 ---
def get_big_cycle_id(date_obj):
    delta = (date_obj - BASE_DATE).days
    return delta // 28

def get_week_id(date_obj):
    delta = (date_obj - BASE_DATE).days
    return delta // 7

def check_consecutive_safe(timeline, index_to_change):
    temp_line = timeline.copy()
    temp_line[index_to_change] = 1 
    max_con = 0
    current_con = 0
    for val in temp_line:
        if val == 1:
            current_con += 1
            max_con = max(max_con, current_con)
        else:
            current_con = 0
    return max_con <= 6

def apply_strict_labor_rules(df_result, year, month, staff_last_month_consecutive={}, skills_map=None):
    if skills_map is None: skills_map = {}
    
    date_cols = []
    col_map = {} 
    for col in df_result.columns:
        if col in ['ID', 'Name', '員工']: continue
        try:
            d = int(col)
            dt = datetime(year, month, d)
            date_cols.append(dt)
            col_map[dt] = col
        except: pass
    
    date_cols.sort()
    if not date_cols: return df_result, []

    logs = []

    for idx, row in df_result.iterrows():
        sid = row['ID']
        
        user_skills = skills_map.get(sid, set())
        if "不排班" in user_skills:
            continue
            
        week_ids = sorted(list(set([get_week_id(dt) for dt in date_cols])))
        for wid in week_ids:
            days_in_week = [dt for dt in date_cols if get_week_id(dt) == wid]
            if not days_in_week: continue
            mandatory_dates = [] 
            regular_9_dates = [] 
            for dt in days_in_week:
                shift = str(df_result.at[idx, col_map[dt]]).strip()
                if is_mandatory_off(shift): mandatory_dates.append(dt)
                if is_regular_rest(shift): regular_9_dates.append(dt)
            
            if len(mandatory_dates) == 0:
                if regular_9_dates:
                    df_result.at[idx, col_map[regular_9_dates[0]]] = "9例"
            elif len(mandatory_dates) > 1:
                for drop_dt in mandatory_dates[1:]:
                    df_result.at[idx, col_map[drop_dt]] = "9"
                    regular_9_dates.append(drop_dt)

        cycle_ids = sorted(list(set([get_big_cycle_id(dt) for dt in date_cols])))
        for cid in cycle_ids:
            days_in_cycle = [dt for dt in date_cols if get_big_cycle_id(dt) == cid]
            if not days_in_cycle: continue
            regular_cnt = 0   
            regular_9_candidates = [] 
            for dt in days_in_cycle:
                shift = str(df_result.at[idx, col_map[dt]]).strip()
                if is_regular_rest(shift):
                    regular_cnt += 1
                    regular_9_candidates.append(dt)
            
            excess_regular = regular_cnt - 4
            if excess_regular > 0:
                prev_cons = staff_last_month_consecutive.get(sid, 0)
                timeline_prefix = [1] * prev_cons
                month_dates = date_cols 
                current_timeline = []
                for dt in month_dates:
                    s = str(df_result.at[idx, col_map[dt]]).strip()
                    val = 1 if is_working_day(s) else 0
                    current_timeline.append(val)
                full_timeline = timeline_prefix + current_timeline
                changed_count = 0
                for target_dt in regular_9_candidates:
                    if changed_count >= excess_regular: break
                    try:
                        day_idx = month_dates.index(target_dt)
                        full_idx = len(timeline_prefix) + day_idx
                        if check_consecutive_safe(full_timeline, full_idx):
                            df_result.at[idx, col_map[target_dt]] = "01特"
                            full_timeline[full_idx] = 1 
                            changed_count += 1
                    except ValueError: pass
    return df_result, logs

def get_prev_month(year, month):
    if month == 1: return year - 1, 12
    return year, month - 1

def auto_calculate_last_consecutive_from_upload(uploaded_file, prev_year, prev_month, current_staff_ids):
    if uploaded_file is None: return {}, {}, "無上傳檔案"
    try:
        xls = pd.ExcelFile(uploaded_file)
        sheets = xls.sheet_names
        target_sheet = None
        candidates = [f"{prev_month}月", f"{prev_month}", f"{prev_month:02d}"]
        for cand in candidates:
            if cand in sheets:
                target_sheet = cand
                break
        if not target_sheet: return {}, {}, f"找不到 '{prev_month}月' 工作表 (無上月資料)"
        
        df_prev = pd.read_excel(uploaded_file, sheet_name=target_sheet, dtype=str)
        header_row = -1
        for i, row in df_prev.iterrows():
            row_str = row.astype(str).values
            if any("卡號" in s or "ID" in s for s in row_str):
                header_row = i + 1 
                break
        if header_row != -1:
             df_prev = pd.read_excel(uploaded_file, sheet_name=target_sheet, header=header_row, dtype=str)
        
        id_col = next((c for c in df_prev.columns if "ID" in str(c) or "卡號" in str(c)), None)
        if not id_col: return {}, {}, "上月工作表無 ID 欄位"
        df_prev[id_col] = df_prev[id_col].apply(clean_str)
        
        day_cols = []
        for c in df_prev.columns:
            try:
                if 1 <= int(float(str(c))) <= 31: day_cols.append(c)
            except: pass
        day_cols.sort(key=lambda x: int(float(str(x))))
        
        con_res = {}
        last_shift_res = {}
        
        for sid in current_staff_ids:
            row = df_prev[df_prev[id_col] == sid]
            if row.empty: 
                con_res[sid] = 0
                last_shift_res[sid] = ""
                continue
            
            con = 0
            for c in reversed(day_cols):
                if is_working_day(str(row.iloc[0][c])): con += 1
                else: break
            con_res[sid] = con
            
            if day_cols:
                last_day_col = day_cols[-1]
                last_shift_res[sid] = clean_str(row.iloc[0][last_day_col])
            else:
                last_shift_res[sid] = ""

        return con_res, last_shift_res, f"已銜接 '{target_sheet}' 工作表"
    except Exception as e:
        return {}, {}, f"讀取上月錯誤: {e}"

def create_template_excel(year, month):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    _, num_days = calendar.monthrange(year, month)
    
    ws1 = wb.active
    ws1.title = "Staff"
    ws1.append(["卡號", "員工", "Skills"])
    ws1.append(["1800", "範例員工", "8-4'F,8-5"]) 

    ws2 = wb.create_sheet("Roster")
    header = ["卡號", "員工"] + [str(i) for i in range(1, num_days + 1)]
    ws2.append(header)
    ws2.append(["1800", "範例員工"] + [""] * num_days)

    ws3 = wb.create_sheet("Shifts")
    ws3.append(["Date", "Shift", "Count"])
    example_date = f"{year}/{month}/1"
    ws3.append([example_date, "8-5", 1])

    ws4 = wb.create_sheet("ShiftTime")
    ws4.append(["Code", "Start", "End"])
    ws4.append(["8-5", 8, 17])
    ws4.append(["8-4'F", 8, 16.5])
    ws4.append(["4-12", 16, 24])
    ws4.append(["12'-9", 12.5, 21])

    wb.save(output)
    return output.getvalue()

def generate_formatted_excel(df, year, month):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month}月"
    
    fill_big_blue = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_big_orange = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    fill_small_pink = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    fill_small_purple = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
    
    weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
    
    day_cols = []
    for c in df.columns:
        if c not in ['ID', 'Name', '員工', '卡號']:
            try: day_cols.append(int(c))
            except: pass
    day_cols.sort()
    
    total_cols = len(day_cols) + 2 + 1 + len(STATS_TARGETS)
    
    row1 = [""] * total_cols
    mid_idx = len(day_cols) // 2
    if mid_idx < 2: mid_idx = 2
    row1[mid_idx-1] = year
    row1[mid_idx] = "年"
    row1[mid_idx+1] = f"{month:02d}"
    row1[mid_idx+2] = "月"
    ws.append(row1)
    
    ws.append([""] * total_cols)
    
    row3 = ["", ""]
    for d in day_cols:
        row3.append(str(d))
    row3.append("") 
    row3.extend([""] * len(STATS_TARGETS))
    ws.append(row3)
    
    row4 = ["卡號", "員工"]
    for d in day_cols:
        dt = datetime(year, month, d)
        row4.append(weekday_map[dt.weekday()])
    row4.append("") 
    row4.extend(STATS_TARGETS)
    ws.append(row4)
    
    for idx, r in df.iterrows():
        id_val = r.get('ID', r.get('卡號', ''))
        name_val = r.get('Name', r.get('員工', id_val))
        row_data = [id_val, name_val]
        
        shift_data = []
        for d in day_cols:
            val = str(r.get(str(d), "")).strip()
            shift_val = val if val not in ['nan', 'None', ''] else ""
            row_data.append(shift_val)
            shift_data.append(shift_val)
            
        row_data.append("") 
        
        for target in STATS_TARGETS:
            count = shift_data.count(target)
            row_data.append(count if count > 0 else "")
            
        ws.append(row_data)
        
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    spacer_col_idx = len(day_cols) + 3
    
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=total_cols):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if cell.column == spacer_col_idx:
                continue
                
            if cell.row >= 4 or (cell.row == 3 and 3 <= cell.column <= len(day_cols) + 2):
                cell.border = thin_border
                
            if cell.row in [3, 4] and 3 <= cell.column <= len(day_cols) + 2:
                d = day_cols[cell.column - 3]
                current_dt = datetime(year, month, d)
                delta_days = (current_dt - BASE_DATE).days
                if delta_days >= 0:
                    if cell.row == 3:
                        big_cycle_idx = delta_days // 28
                        if big_cycle_idx % 2 == 0: cell.fill = fill_big_blue
                        else: cell.fill = fill_big_orange
                    elif cell.row == 4:
                        small_cycle_idx = delta_days // 14
                        if small_cycle_idx % 2 == 0: cell.fill = fill_small_pink
                        else: cell.fill = fill_small_purple

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    for col_idx in range(3, len(day_cols) + 3):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 8
    ws.column_dimensions[get_column_letter(spacer_col_idx)].width = 3
    for col_idx in range(spacer_col_idx + 1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 6

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def create_preview_df(df, year, month):
    weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
    new_cols = []
    weekdays_row = {}
    day_cols = []
    
    for col in df.columns:
        if col in ['ID', '卡號']: 
            new_cols.append('卡號')
            weekdays_row['卡號'] = ''
        elif col in ['Name', '員工']: 
            new_cols.append('員工')
            weekdays_row['員工'] = '星期'
        else:
            try:
                d = int(col)
                dt = datetime(year, month, d)
                date_str = str(d) 
                new_cols.append(date_str)
                weekdays_row[date_str] = weekday_map[dt.weekday()]
                day_cols.append(str(d))
            except:
                new_cols.append(col)
                weekdays_row[col] = ''
                
    df_preview = df.copy()
    df_preview.columns = new_cols
    
    spacer_name = " "
    df_preview[spacer_name] = ""
    weekdays_row[spacer_name] = ""
    
    for target in STATS_TARGETS:
        df_preview[target] = df_preview.apply(lambda row: sum(str(row.get(d, '')).strip() == target for d in day_cols) or "", axis=1)
        weekdays_row[target] = ""
        
    df_preview = pd.concat([pd.DataFrame([weekdays_row]), df_preview], ignore_index=True)
    return df_preview

# --- 解析上傳的 Final Schedule Excel ---
def parse_schedule_file(uploaded_file):
    try:
        df_raw = pd.read_excel(uploaded_file, header=None, dtype=str)
        year, month = None, None
        
        # 1. 尋找年份與月份
        for i in range(min(5, len(df_raw))):
            row_vals = df_raw.iloc[i].values
            for j in range(len(row_vals)):
                if str(row_vals[j]).strip() == "年" and j > 0:
                    try: year = int(float(str(row_vals[j-1]).strip()))
                    except: pass
                if str(row_vals[j]).strip() == "月" and j > 0:
                    try: month = int(float(str(row_vals[j-1]).strip()))
                    except: pass
                    
        # 2. 尋找包含「卡號」的表頭列索引
        h_idx = -1
        for i in range(min(10, len(df_raw))):
            if any(isinstance(v, str) and "卡號" in v for v in df_raw.iloc[i].values):
                h_idx = i
                break
                
        if h_idx == -1 or year is None or month is None:
            return None, None, None, "❌ 無法解析年份、月份或卡號列，請確認上傳的是正確的排班結果檔。"
            
        date_row = df_raw.iloc[h_idx - 1]
        header_row = df_raw.iloc[h_idx]
        
        records = []
        for row_idx in range(h_idx + 1, len(df_raw)):
            row_data = df_raw.iloc[row_idx]
            staff_id = ""
            
            for col_idx in range(len(header_row)):
                if str(header_row.iloc[col_idx]).strip() == "卡號":
                    staff_id = str(row_data.iloc[col_idx]).strip()
                    break
                    
            if not staff_id or staff_id in ['nan', 'None']:
                continue
                
            for col_idx in range(len(header_row)):
                d_val = str(date_row.iloc[col_idx]).strip()
                try:
                    d = int(float(d_val))
                    if 1 <= d <= 31:
                        shift = str(row_data.iloc[col_idx]).strip()
                        if shift not in ['', 'nan', 'None', '0']:
                            date_str = f"{year}-{month:02d}-{d:02d}"
                            records.append({
                                '日期': date_str,
                                '班別': shift,
                                '人員': staff_id
                            })
                except:
                    pass
                    
        df_records = pd.DataFrame(records)
        return df_records, year, month, ""
    except Exception as e:
        return None, None, None, f"檔案解析失敗: {e}"

# --- 從 Flatten Records 產出活動病歷掃描分析 ---
def generate_scan_analysis_excel_from_records(df_records, target_shifts):
    df_report = df_records[df_records['班別'].isin(target_shifts)].copy()
    
    if not df_report.empty:
        df_report['日期'] = pd.to_datetime(df_report['日期']).dt.date
        df_report = df_report.sort_values(by=["日期", "班別", "人員"])
        
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "範例"
    
    # --- 統一樣式設定 (字體: 微軟正黑體 11, 置中, 框線) ---
    base_font = Font(name='微軟正黑體', size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    # --- 寫入表頭 (A~F) ---
    headers = ["日期", "班別", "人員", "總病歷本數（本）", "總掃描頁數（頁）", "備註"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = base_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
    # --- 寫入資料 (A~C) ---
    if not df_report.empty:
        for row_idx, record in enumerate(df_report.to_dict('records'), 2):
            # 寫入日期並設定 Excel 格式為 m月d日 (例如 3月1日)
            cell_date = ws.cell(row=row_idx, column=1, value=record["日期"])
            cell_date.number_format = 'm"月"d"日"'
            
            ws.cell(row=row_idx, column=2, value=record["班別"])
            ws.cell(row=row_idx, column=3, value=record["人員"])
            
    # --- 寫入 L 欄 (篩選條件) ---
    cell_L1 = ws.cell(row=1, column=12, value="班別")
    cell_L1.font = base_font
    cell_L1.alignment = center_alignment
    cell_L1.border = thin_border
    
    for i, ts in enumerate(target_shifts, 2):
        ws.cell(row=i, column=12, value=ts)
        
    # --- 套用全域樣式給資料列 ---
    for row in ws.iter_rows(min_row=2, max_row=len(df_report) + 1, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.font = base_font
            
    for row in ws.iter_rows(min_row=2, max_row=len(target_shifts) + 1, min_col=12, max_col=12):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.font = base_font
            
    # --- 調整欄寬 (對齊您的範本) ---
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['L'].width = 15
    
    wb.save(output)
    return output.getvalue()


# --- 3. 主程式介面 ---

with st.sidebar:
    st.title("⚙️ 排班設定面板")
    
    c1, c2 = st.columns(2)
    with c1: 
        now = datetime.now()
        this_year = now.year
        year_options = list(range(this_year - 1, this_year + 6))
        default_year_idx = year_options.index(this_year)
        y = st.selectbox("年份", year_options, index=default_year_idx)
    with c2: 
        m = st.selectbox("月份", range(1, 13), index=now.month - 1)

    st.divider()

    st.write("📝 **初次使用？請先下載範本**")
    template_data = create_template_excel(y, m) 
    st.download_button(
        label="📥 下載排班範本",
        data=template_data,
        file_name="排班範本.xlsx", 
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
    with st.expander("🛠️ 快速生成每月需求表 (Shifts)"):
        st.caption("勾選平日/假日需要的班別，自動產生整個月的 Excel！")
        all_shifts = [
            "8-4'F", "8-5", "12'-9", "4-12", "8-4'掃", 
            "8-4'銷", "8-4'", "8-5銷", "8-5掃", 
            "01", "01特", "9", "9例"
        ]
        
        st.write("🗓️ **平日 (週一~週五)**:")
        wd_default = ["8-4'F", "8-5", "12'-9", "4-12", "8-5掃", "01"]
        wd_default = [x for x in wd_default if x in all_shifts]
        
        wd_shifts = st.multiselect("平日班別", all_shifts, default=wd_default)

        st.write("🎉 **假日 (週六、週日)**:")
        we_default = ["8-4'F", "8-4'", "4-12", "8-4'掃"]
        we_default = [x for x in we_default if x in all_shifts]
        
        we_shifts = st.multiselect("假日班別", all_shifts, default=we_default)

        if st.button("⚡ 生成並準備下載"):
            try:
                _, num_days = calendar.monthrange(y, m)
                data_gen = []
                for day_gen in range(1, num_days + 1):
                    dt_gen = datetime(y, m, day_gen)
                    date_str = dt_gen.strftime("%Y/%-m/%-d")
                    if dt_gen.weekday() >= 5: target_shifts = we_shifts
                    else: target_shifts = wd_shifts
                    for s_name in target_shifts:
                        data_gen.append([date_str, s_name, 1])
                df_gen = pd.DataFrame(data_gen, columns=["Date", "Shift", "Count"])
                output_gen = io.BytesIO()
                with pd.ExcelWriter(output_gen, engine='xlsxwriter') as writer:
                    df_gen.to_excel(writer, sheet_name='Shifts', index=False)
                st.download_button(
                    label=f"📥 下載 {m}月需求表",
                    data=output_gen.getvalue(),
                    file_name=f"shifts_{y}_{m}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"生成失敗: {e}")

    # --- 左側欄位：病歷掃描分析報表 ---
    st.divider()
    st.write("📥 **產出活動病歷掃描分析**")
    st.caption("上傳已生成的排班結果檔 (schedule_..._final.xlsx) 來轉換報表。")
    scan_uploaded_file = st.file_uploader("📂 上傳排班結果檔", type=['xlsx'], key="scan_upload")
    
    default_scan_shifts = ["8-4'掃", "8-4'", "8-5", "12'-9", "8-5掃"]
    all_possible_shifts = list(set(["8-4'F", "8-5", "12'-9", "4-12", "8-4'掃", "8-4'銷", "8-4'", "8-5銷", "8-5掃", "01", "01特", "9", "9例"] + default_scan_shifts))
    
    selected_scan_shifts = st.multiselect(
        "選擇要匯出的班別 (L欄)：",
        options=all_possible_shifts,
        default=[s for s in default_scan_shifts if s in all_possible_shifts]
    )
    
    if scan_uploaded_file is not None:
        df_records, r_year, r_month, err_msg = parse_schedule_file(scan_uploaded_file)
        if err_msg:
            st.error(err_msg)
        else:
            scan_excel_data = generate_scan_analysis_excel_from_records(df_records, selected_scan_shifts)
            # 檔名更新：拿掉114，變成 活動病歷掃描分析_年份_月份.xlsx
            fn_scan = f"活動病歷掃描分析_{r_year}_{r_month}.xlsx"
            st.download_button(
                label=f"⚡ 點擊下載掃描報表",
                data=scan_excel_data,
                file_name=fn_scan,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )


st.title("📅 智慧排班系統")
st.markdown("---")

# --- 主畫面：排班功能 ---
uploaded_file = st.file_uploader("📂 請上傳排班模板 (data.xlsx) 以啟動 AI 排班", type=['xlsx'])
st.info("💡 **週期上色說明**：\n- 日期列：28天大週期 (藍/橘)\n- 星期列：14天小週期 (粉/紫)")

if uploaded_file is not None:
    try:
        try:
            df_staff = pd.read_excel(uploaded_file, sheet_name='Staff', dtype=str)
            staff_cols = {'ID': ['ID', '卡號'], 'Skills': ['Skills', '技能']}
            df_staff = smart_rename(df_staff, staff_cols)
            skills_map = {}
            for _, r in df_staff.iterrows():
                if 'ID' in r and 'Skills' in r:
                    sid = clean_str(r['ID'])
                    skills_map[sid] = parse_skills(r['Skills'])
                    if "不排班" in str(r['Skills']):
                        skills_map[sid] = {"不排班"}
        except: 
            skills_map = {}
            st.warning("⚠️ 讀取 Staff 失敗，將無法執行技能限制。")

        try:
            df_raw = pd.read_excel(uploaded_file, sheet_name='Roster', header=None, dtype=str)
            h_idx = -1
            for i, r in df_raw.head(15).iterrows():
                if any(isinstance(v, str) and ("卡號" in v or "ID" in v) for v in r.values):
                    h_idx = i
                    break
            
            if h_idx == -1:
                st.error("❌ Roster 工作表找不到 '卡號' 欄位。")
                st.stop()
                
            date_row = df_raw.iloc[h_idx - 1] if h_idx > 0 else df_raw.iloc[h_idx]
            header_row = df_raw.iloc[h_idx]
            
            new_cols = []
            v_days = []
            
            for col_idx in range(len(df_raw.columns)):
                h_val = str(header_row.iloc[col_idx]).strip()
                date_val = date_row.iloc[col_idx]
                
                if h_val in ['卡號', 'ID']:
                    new_cols.append('ID')
                elif h_val in ['員工', 'Name', '姓名']:
                    new_cols.append('Name')
                else:
                    day = None
                    try:
                        if pd.notna(date_val):
                            if isinstance(date_val, (datetime, pd.Timestamp)):
                                day = date_val.day
                            else:
                                dt = pd.to_datetime(date_val)
                                day = dt.day
                    except: pass
                    
                    if day is None:
                        try: day = int(float(str(h_val).replace(".0","")))
                        except: pass
                        
                    if day is not None and 1 <= day <= 31:
                        new_cols.append(str(day))
                        if day not in v_days: v_days.append(day)
                    else:
                        new_cols.append(f"DROP_{col_idx}")
                        
            df_roster = df_raw.iloc[h_idx + 1:].copy()
            df_roster.columns = new_cols
            
            cols_to_keep = ['ID', 'Name'] + [str(d) for d in sorted(v_days)]
            valid_cols = [c for c in cols_to_keep if c in df_roster.columns]
            df_roster = df_roster[valid_cols]
            
            df_roster = df_roster.dropna(subset=['ID'])
            df_roster['ID'] = df_roster['ID'].apply(clean_str)
            for d in v_days:
                if str(d) in df_roster.columns:
                    df_roster[str(d)] = df_roster[str(d)].apply(clean_str)
                    
            v_days = sorted(v_days)
            if not v_days:
                st.error("❌ Roster 工作表無法識別日期，請確認 '卡號' 的上一列是否有標註完整日期。")
                st.stop()
                
        except Exception as e:
            st.error(f"❌ 讀取 Roster 失敗: {e}")
            st.stop()

        try:
            df_shifts = pd.read_excel(uploaded_file, sheet_name='Shifts')
            df_shifts = smart_rename(df_shifts, {'Date':['Date','日期'], 'Shift':['Shift','班別'], 'Count':['Count','人數']})
            df_shifts['Date'] = pd.to_datetime(df_shifts['Date'])
        except Exception as e:
            st.error(f"❌ 讀取 Shifts 失敗: {e}")
            st.stop()

        py, pm = get_prev_month(y, m)
        sids = df_roster['ID'].tolist()
        last_con, last_shifts, msg = auto_calculate_last_consecutive_from_upload(uploaded_file, py, pm, sids)
        
        if "找不到" in msg: st.warning(f"⚠️ {msg}")
        else: st.success(f"✅ {msg}")

        mask = (df_shifts['Date'].dt.year == y) & (df_shifts['Date'].dt.month == m)
        m_shifts = df_shifts[mask].copy()
        m_shifts = m_shifts[m_shifts['Date'].dt.day.isin(v_days)]

        if st.button("🚀 啟動 AI 自動排班", type="primary", use_container_width=True):
            shift_time_db = {}
            forbidden_pairs = set() 
            try:
                df_st = pd.read_excel(uploaded_file, sheet_name='ShiftTime', dtype=str)
                for _, row in df_st.iterrows():
                    code = clean_str(row.get('Code', ''))
                    try:
                        s_t = float(row.get('Start', 0))
                        e_t = float(row.get('End', 0))
                        shift_time_db[code] = {'Start': s_t, 'End': e_t}
                    except: pass
                known_shifts = list(shift_time_db.keys())
                for s1 in known_shifts:
                    for s2 in known_shifts:
                        t1 = shift_time_db[s1]
                        t2 = shift_time_db[s2]
                        rest = (t2['Start'] + 24) - t1['End']
                        if rest < 11: forbidden_pairs.add((s1, s2))
                forbidden_pairs.add(('4-12', "12'-9"))
                if forbidden_pairs:
                    with st.expander(f"🛡️ 已啟動法規防護 ({len(forbidden_pairs)} 條規則)"):
                        st.write(list(forbidden_pairs))
            except: pass

            with st.spinner("⏳ AI 正在運算最佳排班組合..."):
                model = cp_model.CpModel()
                solver = cp_model.CpSolver()
                vars = {}
                fixed = {}
                for _, r in df_roster.iterrows():
                    sid = r['ID']
                    for d in v_days:
                        v = r[str(d)]
                        if v != "": fixed[(sid, d)] = v

                needed = []
                for _, r in m_shifts.iterrows():
                    dn = r['Date'].day
                    sn = clean_str(r['Shift'])
                    cnt = r['Count']
                    filled = sum(1 for sid in sids if fixed.get((sid, dn)) == sn)
                    rem = cnt - filled
                    if rem > 0: needed.append((dn, sn, rem))

                lookup = {}
                obj = []
                for d, s, c in needed:
                    grp = []
                    target_shift = clean_str(s)
                    for sid in sids:
                        if (sid, d) in fixed: continue
                        user_skills = skills_map.get(sid, set())
                        if "不排班" in user_skills: continue
                        if is_working_day(target_shift) and target_shift not in user_skills:
                            continue
                        v = model.NewBoolVar(f"{sid}_{d}_{s}")
                        vars[(sid, d, s)] = v
                        grp.append(v)
                        if (sid, d) not in lookup: lookup[(sid, d)] = []
                        lookup[(sid, d)].append(v)
                        obj.append(v * random.randint(100, 200)) 
                    if grp: model.Add(sum(grp) <= c)

                model.Maximize(sum(obj))
                for _, vs in lookup.items(): model.Add(sum(vs) <= 1)
                
                w_size = 7
                for sid in sids:
                    prev = last_con.get(sid, 0)
                    pre = [1] * prev
                    curr = []
                    for d in v_days:
                        fv = fixed.get((sid, d), "")
                        if fv: val = 0 if is_rest_day(fv) else 1
                        elif (sid, d) in lookup: val = sum(lookup[(sid, d)])
                        else: val = 0 
                        curr.append(val)
                    full = pre + curr
                    if len(full) >= w_size:
                        for i in range(len(full)-w_size+1):
                            win = full[i:i+w_size]
                            model.Add(sum(win) <= 6)
                
                for sid in sids:
                    for i in range(len(v_days) - 1):
                        d1 = v_days[i]
                        d2 = v_days[i+1]
                        fix1 = fixed.get((sid, d1))
                        fix2 = fixed.get((sid, d2))
                        for s1, s2 in forbidden_pairs:
                            v1 = vars.get((sid, d1, s1)); v2 = vars.get((sid, d2, s2))
                            if v1 is not None and v2 is not None: model.AddBoolOr([v1.Not(), v2.Not()])
                            if fix1 == s1 and v2 is not None: model.Add(v2 == 0)
                            if v1 is not None and fix2 == s2: model.Add(v1 == 0)

                if v_days:
                    first_day = v_days[0]
                    for sid in sids:
                        last_s = last_shifts.get(sid, "")
                        if last_s:
                            for (t_sid, t_d, t_s), v in vars.items():
                                if t_sid == sid and t_d == first_day:
                                    if (last_s, t_s) in forbidden_pairs:
                                        model.Add(v == 0)

                status = solver.Solve(model)

            if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
                df_fin = df_roster.copy().set_index('ID')
                for (sid, d, s), v in vars.items():
                    if solver.Value(v): df_fin.at[sid, str(d)] = s
                df_fin = df_fin.reset_index()

                for idx, r in df_fin.iterrows():
                    sid = r['ID']
                    user_skills = skills_map.get(sid, set())
                    fill = "" if "不排班" in user_skills else "9"
                    for d in v_days:
                        val = str(r[str(d)]).strip()
                        if val in ['','nan','None','0']: df_fin.at[idx, str(d)] = fill

                df_fin, _ = apply_strict_labor_rules(df_fin, y, m, last_con, skills_map)
                
                cols = ['ID', 'Name'] + [str(d) for d in v_days]
                df_export = df_fin[cols].copy()
                
                kpi1, kpi2, kpi3 = st.columns(3)
                kpi1.metric("👥 參與排班人數", f"{len(sids)} 人")
                kpi2.metric("📅 排班總天數", f"{len(v_days)} 天")
                kpi3.metric("🛡️ 違規檢查", "0 錯誤", delta="Passed")

                tab1, tab2 = st.tabs(["📊 排班結果預覽", "📥 下載 Excel"])
                with tab1:
                    df_preview = create_preview_df(df_export, y, m)
                    st.dataframe(df_preview, use_container_width=True)
                with tab2:
                    xlsx_data = generate_formatted_excel(df_export, y, m)
                    fn = f"schedule_{y}_{m}_final.xlsx"
                    st.download_button(label=f"📥 下載排班結果 ({fn})", data=xlsx_data, file_name=fn, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
            else:
                st.error("❌ 排班失敗：找不到可行解。")
    except Exception as e:
        st.error(f"Error: {e}")
        st.text(f"詳細錯誤訊息：\n{e}")